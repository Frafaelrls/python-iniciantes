""" Add a valuer in a sheet
For complement exercises add a name of fifth column. Done.
Format a fist cell from the fifth column. Not Done."""
import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}  # dictionary

"""This part is additional, not asked in training

name_for_column = product_list.cell(1, 5)  # access row 1 column 5
name_for_column.value = 'Total Value'  # writing in cell

For this exemple, every time when same function is used before a equal this
 function is for save a data, when is used after is used to access a data"""

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value  # access column 2
    price = product_list.cell(product_row, 3).value  # access column 3
    product_number = product_list.cell(product_row, 1).value  # access column 1
    inventory_price = product_list.cell(product_row, 5)  # access column 5
    # calculation a number of products per supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = \
            current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic for products inventory less than 10
    if inventory < 10:
        product_under_10_inv[int(product_number)] = int(inventory)

    # add value fot total inventory price
    inventory_price.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")  # save in a new file

print(product_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)
