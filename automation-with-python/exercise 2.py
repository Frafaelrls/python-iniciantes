""" Calculate a value for with company"""
import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}  # Dictionary

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value  # access column 2
    price = product_list.cell(product_row, 3).value  # access column 3
    # calculation a number of products per supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding a new supplier")
        product_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:  # verify in dictionary if
        # already exist and do
        current_total_value = total_value_per_supplier.get(supplier_name)  #
        # get a current value from dictionary
        total_value_per_supplier[supplier_name] = \
            current_total_value + inventory * price  # add the new value
        # to the old value
    else:  # for fist time add a new value
        total_value_per_supplier[supplier_name] = inventory * price  #
        # calculate total value

print(product_per_supplier)
print(total_value_per_supplier)
