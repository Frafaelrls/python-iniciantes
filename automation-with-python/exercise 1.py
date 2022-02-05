"""How many product have in supplier
List a supplier name and his products."""
import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")  # Open a file
product_list = inv_file["Sheet1"]  # Read a specific sheet

products_per_supplier = {}  # Dictionary

# print(product_list.max_row)
for product_row in range(2, product_list.max_row + 1):  # Start and finish
    supplier_name = product_list.cell(product_row, 4).value  # Save date cell

    if supplier_name in products_per_supplier:  # Verif in dictionary if
        # already exists do
        current_num_products = products_per_supplier.get(supplier_name)  # get
        # value from 'key'
        products_per_supplier[supplier_name] = current_num_products + 1  # add
        # 1 in existing supplier
    else:  # For new supplier
        print("Adding a new supplier")
        products_per_supplier[supplier_name] = 1  # Create new data in
        # dictionary and set as one

print(products_per_supplier)
